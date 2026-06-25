"""CLI for OCR/parsing/indexing the local 정리 study folder."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

from dotenv import load_dotenv

from exam_material_ingestion import SUBJECT_KEYWORDS, summarize_ingestion_results
from langchain_service import LangChainService
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter


DEFAULT_STUDY_DIR = Path("정리")
SUPPORTED_PDF_EXTENSIONS = {".pdf"}
SUPPORTED_EXCEL_EXTENSIONS = {".xlsm", ".xlsx"}
SUPPORTED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg"}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=DEFAULT_STUDY_DIR)
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--limit", type=int)
    parser.add_argument("--supplemental-only", action="store_true")
    parser.add_argument("--retag-existing", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--report", type=Path, default=Path("study_folder_ingest_report.json"))
    args = parser.parse_args()

    load_dotenv()
    root = args.root.expanduser()
    pdfs = list_study_pdfs(root)
    excels = list_excel_materials(root)
    images = list_image_materials(root)
    supplemental = [*excels, *images]
    offset = max(0, args.offset)
    selected = pdfs[offset:]
    if args.limit is not None:
        selected = selected[: max(0, args.limit)]
    if args.supplemental_only:
        selected = []

    splitter_config = study_splitter_config()
    if args.retag_existing:
        service = LangChainService()
        report = retag_existing_study_documents(service, root)
        write_report(args.report, report)
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return 0 if report["failed_count"] == 0 else 1

    if args.dry_run:
        report = {
            "status": "dry_run",
            "root": str(root),
            "pdf_files_count": len(pdfs),
            "selected_count": len(selected),
            "excel_files_count": len(excels),
            "image_files_count": len(images),
            "supplemental_files_count": len(supplemental),
            "offset": offset,
            "splitter": splitter_config,
            "preview": [
                {
                    "filename": pdf.name,
                    "relative_path": str(pdf.relative_to(root)),
                    "metadata": infer_study_metadata(pdf, root),
                }
                for pdf in selected[:80]
            ],
            "supplemental_preview": [
                {
                    "filename": path.name,
                    "relative_path": str(path.relative_to(root)),
                    "suffix": path.suffix.lower(),
                    "metadata": infer_study_metadata(path, root),
                }
                for path in supplemental[:80]
            ],
        }
        write_report(args.report, report)
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return 0

    service = LangChainService()
    results: list[dict[str, Any]] = []
    supplemental_results: list[dict[str, Any]] = []
    for index, pdf in enumerate(selected, start=offset + 1):
        metadata = infer_study_metadata(pdf, root)
        result = service.process_pdf(
            str(pdf),
            metadata_override=metadata,
            **splitter_config,
        )
        row = {
            "sequence": index,
            "filename": pdf.name,
            "relative_path": str(pdf.relative_to(root)),
            **metadata,
            **result,
        }
        results.append(row)
        print(
            f"[{index}/{len(pdfs)}] {result.get('status')} "
            f"{pdf.name} chunks={result.get('chunks_count', 0)} "
            f"indexed={result.get('indexed')}"
        )
        write_report(
            args.report,
            build_report(
                root,
                offset,
                pdfs,
                excels,
                images,
                splitter_config,
                results,
                supplemental_results,
                complete=False,
            ),
        )

    for index, path in enumerate(supplemental, start=1):
        metadata = infer_study_metadata(path, root)
        result = process_supplemental_material(service, path, metadata, splitter_config)
        row = {
            "sequence": index,
            "filename": path.name,
            "relative_path": str(path.relative_to(root)),
            **metadata,
            **result,
        }
        supplemental_results.append(row)
        print(
            f"[supplemental {index}/{len(supplemental)}] {result.get('status')} "
            f"{path.name} chunks={result.get('chunks_count', 0)} "
            f"indexed={result.get('indexed')}"
        )
        write_report(
            args.report,
            build_report(
                root,
                offset,
                pdfs,
                excels,
                images,
                splitter_config,
                results,
                supplemental_results,
                complete=False,
            ),
        )

    report = build_report(
        root,
        offset,
        pdfs,
        excels,
        images,
        splitter_config,
        results,
        supplemental_results,
        complete=True,
    )
    write_report(args.report, report)
    print(
        json.dumps(
            {
                key: value
                for key, value in report.items()
                if key not in {"results", "supplemental_results"}
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0 if report["failed_count"] == 0 and report["supplemental_failed_count"] == 0 else 1


def list_study_pdfs(root: str | Path = DEFAULT_STUDY_DIR) -> list[Path]:
    base = Path(root).expanduser()
    if not base.exists():
        return []
    return sorted(
        path
        for path in base.rglob("*")
        if path.is_file()
        and not path.name.startswith(".")
        and path.suffix.lower() in SUPPORTED_PDF_EXTENSIONS
    )


def list_excel_materials(root: str | Path = DEFAULT_STUDY_DIR) -> list[Path]:
    base = Path(root).expanduser()
    if not base.exists():
        return []
    return sorted(
        path
        for path in base.rglob("*")
        if path.is_file()
        and not path.name.startswith(".")
        and path.suffix.lower() in SUPPORTED_EXCEL_EXTENSIONS
    )


def list_image_materials(root: str | Path = DEFAULT_STUDY_DIR) -> list[Path]:
    base = Path(root).expanduser()
    if not base.exists():
        return []
    return sorted(
        path
        for path in base.rglob("*")
        if path.is_file()
        and not path.name.startswith(".")
        and path.suffix.lower() in SUPPORTED_IMAGE_EXTENSIONS
    )


def infer_study_metadata(path: str | Path, root: str | Path = DEFAULT_STUDY_DIR) -> dict[str, Any]:
    file_path = Path(path)
    base = Path(root).expanduser()
    try:
        relative = file_path.relative_to(base)
    except ValueError:
        relative = Path(file_path.name)
    relative_text = normalize(str(relative))
    search_text = normalize(f"{relative_text} {file_path.stem}")
    category_code, category_name = study_category(relative)
    subject = subject_from_text(search_text)
    role = material_role(search_text)
    application = application_profile(category_code, role, search_text)
    return {
        "collection": "exam_application",
        "source_collection": "study_folder",
        "knowledge_layer": "exam_application",
        "document_type": document_type(category_code, role),
        "study_category_code": category_code,
        "study_category": category_name,
        "material_role": role,
        **application,
        "subject": subject,
        "source_folder": str(base),
        "relative_path": relative_text,
        "exam_domain": "교직논술" if "논술" in search_text else "교육과정",
        "study_priority": study_priority(category_code, role),
    }


def study_splitter_config() -> dict[str, Any]:
    return {
        "chunk_size": 450,
        "chunk_overlap": 90,
        "separators": [
            "\n\n",
            "\n[",
            "\n※",
            "\n<",
            "\n문제",
            "\n문항",
            "\n정답",
            "\n답안",
            "\n해설",
            "\n성취기준",
            "\n",
            ". ",
            " ",
            "",
        ],
    }


def build_report(
    root: Path,
    offset: int,
    pdfs: list[Path],
    excels: list[Path],
    images: list[Path],
    splitter_config: dict[str, Any],
    results: list[dict[str, Any]],
    supplemental_results: list[dict[str, Any]],
    complete: bool,
) -> dict[str, Any]:
    supplemental_failed = sum(result.get("status") != "success" for result in supplemental_results)
    supplemental_indexed = sum(result.get("indexed") is True for result in supplemental_results)
    supplemental_skipped = sum(
        result.get("status") == "success" and result.get("indexed") is False
        for result in supplemental_results
    )
    supplemental_chunks = sum(int(result.get("chunks_count") or 0) for result in supplemental_results)
    return {
        **summarize_ingestion_results(results),
        "root": str(root),
        "offset": offset,
        "pdf_files_count": len(pdfs),
        "processed_count": len(results),
        "excel_files_count": len(excels),
        "image_files_count": len(images),
        "supplemental_files_count": len(excels) + len(images),
        "supplemental_processed_count": len(supplemental_results),
        "supplemental_indexed_count": supplemental_indexed,
        "supplemental_skipped_count": supplemental_skipped,
        "supplemental_failed_count": supplemental_failed,
        "supplemental_chunks_count": supplemental_chunks,
        "splitter": splitter_config,
        "complete": complete,
        "results": results,
        "supplemental_results": supplemental_results,
    }


def retag_existing_study_documents(service: LangChainService, root: Path) -> dict[str, Any]:
    if service.vector_store is None:
        return {
            "status": "error",
            "root": str(root),
            "retagged_chunks": 0,
            "failed_count": 1,
            "message": "FAISS vector store를 찾지 못했습니다.",
        }
    docstore = getattr(getattr(service.vector_store, "docstore", None), "_dict", {})
    changed = 0
    failed = 0
    category_counts: dict[str, int] = {}
    use_counts: dict[str, int] = {}
    for document in docstore.values():
        metadata = document.metadata
        source = metadata.get("source") or ""
        relative = metadata.get("relative_path")
        if not is_study_folder_document(metadata, source):
            continue
        try:
            path = Path(source)
            if relative:
                relative_path = Path(relative)
            else:
                try:
                    relative_path = path.relative_to(root)
                except ValueError:
                    relative_path = Path(path.name)
            inferred = infer_study_metadata(path, root)
            metadata.update({
                key: value
                for key, value in inferred.items()
                if value is not None
            })
            metadata["relative_path"] = normalize(str(relative_path))
            metadata["retagged_at"] = "2026-06-25"
            category = metadata.get("study_category") or "unknown"
            use = metadata.get("exam_application_use") or "unknown"
            category_counts[category] = category_counts.get(category, 0) + 1
            use_counts[use] = use_counts.get(use, 0) + 1
            changed += 1
        except Exception:
            failed += 1
    if changed:
        service.vector_store.save_local(
            folder_path=service.faiss_db_path,
            index_name="index",
        )
    return {
        "status": "success" if failed == 0 else "partial_success",
        "root": str(root),
        "retagged_chunks": changed,
        "failed_count": failed,
        "collection": "exam_application",
        "knowledge_layer": "exam_application",
        "category_counts": category_counts,
        "exam_application_use_counts": use_counts,
    }


def is_study_folder_document(metadata: dict[str, Any], source: str) -> bool:
    return (
        metadata.get("collection") in {"study_folder", "exam_application"}
        or metadata.get("source_collection") == "study_folder"
        or "정리/" in source
        or "정리/" in source
    )


def process_supplemental_material(
    service: LangChainService,
    path: Path,
    metadata: dict[str, Any],
    splitter_config: dict[str, Any],
) -> dict[str, Any]:
    try:
        suffix = path.suffix.lower()
        if suffix in SUPPORTED_EXCEL_EXTENSIONS:
            text, warnings = extract_excel_text(path)
            method = "openpyxl"
        elif suffix in SUPPORTED_IMAGE_EXTENSIONS:
            text, warnings = extract_image_text(path)
            method = "image_ocr"
        else:
            return {"status": "skipped", "message": f"지원하지 않는 파일 형식: {suffix}", "indexed": False}
        return index_text_material(
            service=service,
            path=path,
            text=text,
            metadata={**metadata, "document_type": f"{metadata.get('document_type')}_supplemental"},
            extraction_method=method,
            warnings=warnings,
            splitter_config=splitter_config,
        )
    except Exception as exc:
        return {"status": "error", "message": str(exc), "indexed": False}


def extract_excel_text(path: Path) -> tuple[str, list[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    lines: list[str] = [f"# {path.name}"]
    warnings: list[str] = []
    try:
        for sheet in workbook.worksheets:
            lines.append(f"\n## Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [normalize_cell(value) for value in row]
                values = [value for value in values if value]
                if values:
                    lines.append(" | ".join(values))
    finally:
        workbook.close()
    text = "\n".join(lines).strip()
    if len(text) < 30:
        warnings.append("엑셀에서 추출한 텍스트가 매우 적습니다.")
    return text, warnings


def extract_image_text(path: Path) -> tuple[str, list[str]]:
    import pytesseract
    from PIL import Image

    warnings: list[str] = []
    image = Image.open(path)
    try:
        available = set(pytesseract.get_languages(config=""))
        language = "kor+eng" if "kor" in available else "eng"
        text = pytesseract.image_to_string(image, lang=language, config="--oem 3 --psm 6")
    finally:
        image.close()
    text = normalize(text).strip()
    if len(re.sub(r"\s", "", text)) < 20:
        warnings.append("이미지 OCR 텍스트가 매우 적습니다.")
    return f"# {path.name}\n\n{text}", warnings


def index_text_material(
    service: LangChainService,
    path: Path,
    text: str,
    metadata: dict[str, Any],
    extraction_method: str,
    warnings: list[str],
    splitter_config: dict[str, Any],
) -> dict[str, Any]:
    normalized_text = normalize(text).strip()
    if not normalized_text:
        raise ValueError("색인 가능한 텍스트를 추출하지 못했습니다")
    digest = hashlib.sha256()
    digest.update(str(path).encode("utf-8"))
    digest.update(normalized_text.encode("utf-8"))
    document_id = digest.hexdigest()
    if document_id in service.indexed_document_ids:
        return {
            "status": "success",
            "message": "이미 색인된 동일 자료입니다.",
            "chunks_count": 0,
            "pages_count": 1,
            "document_id": document_id,
            "extraction_methods": [extraction_method],
            "warnings": warnings,
            "indexed": False,
        }
    base_document = Document(
        page_content=normalized_text,
        metadata={
            **metadata,
            "document_id": document_id,
            "source": str(path),
            "page": 0,
            "page_number": 1,
            "extraction_method": extraction_method,
        },
    )
    splitter = RecursiveCharacterTextSplitter(**splitter_config)
    chunks = splitter.split_documents([base_document])
    for index, chunk in enumerate(chunks, start=1):
        chunk.metadata["chunk_number"] = index
        chunk.metadata["chunk_size"] = splitter_config["chunk_size"]
        chunk.metadata["chunk_overlap"] = splitter_config["chunk_overlap"]
    service._index_chunks(chunks, document_id)
    return {
        "status": "success",
        "message": f"보조 자료 처리 완료: {len(chunks)}개 청크 생성",
        "chunks_count": len(chunks),
        "pages_count": 1,
        "document_id": document_id,
        "extraction_methods": [extraction_method],
        "warnings": warnings,
        "indexed": True,
        "document_type": metadata.get("document_type"),
    }


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return normalize(str(value)).strip()


def write_report(path: Path, data: dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def study_category(relative: Path) -> tuple[str | None, str | None]:
    first = normalize(relative.parts[0]) if relative.parts else ""
    match = re.match(r"\[(\d{2})\]\[([^\]]+)\]", first)
    if match:
        return match.group(1), match.group(2)
    match = re.match(r"\[(\d{2})\]", first)
    if match:
        return match.group(1), first
    return None, first or None


def subject_from_text(text: str) -> str | None:
    for keyword, subject in SUBJECT_KEYWORDS.items():
        if keyword in text:
            return subject
    return None


def material_role(text: str) -> str:
    if any(token in text for token in ["해설", "정답", "답안", "오답"]):
        return "explanation"
    if any(token in text for token in ["문제", "연문", "연습"]):
        return "question"
    if any(token in text for token in ["기출", "분석"]):
        return "analysis"
    if any(token in text for token in ["원문", "교육과정"]):
        return "reference"
    if any(token in text for token in ["정리", "스제트", "구조화", "암기", "핵심"]):
        return "summary"
    return "reference"


def document_type(category_code: str | None, role: str) -> str:
    category_map = {
        "00": "exam_application_practice",
        "01": "exam_application_subject_theory",
        "02": "exam_application_basic_theory",
        "03": "exam_application_curriculum",
        "04": "exam_application_past_exam_analysis",
        "05": "exam_application_essay",
        "08": "exam_application_exam_day",
        "09": "exam_application_output_table",
        "10": "exam_application_general_curriculum",
    }
    base = category_map.get(category_code, "exam_application_material")
    return f"{base}_{role}" if role not in {"reference", "summary"} else base


def application_profile(category_code: str | None, role: str, text: str) -> dict[str, Any]:
    if role == "question":
        use = "practice_question"
    elif role == "explanation":
        use = "answer_explanation"
    elif role == "analysis":
        use = "past_exam_pattern"
    elif role == "summary":
        use = "recall_output"
    else:
        use = "exam_reference"
    if category_code == "04":
        use = "past_exam_pattern"
    elif category_code == "05":
        use = "essay_application"
    elif category_code == "08":
        use = "exam_day_recall"
    elif category_code == "09":
        use = "output_table"

    axis = "curriculum_to_exam"
    if any(token in text for token in ["모형", "단계"]):
        axis = "teaching_model_output"
    elif any(token in text for token in ["성취기준", "내용체계"]):
        axis = "standard_to_item"
    elif any(token in text for token in ["답안", "해설", "정답"]):
        axis = "answer_basis"
    elif any(token in text for token in ["기출", "분석"]):
        axis = "past_exam_pattern"
    elif "논술" in text:
        axis = "essay_answer"

    return {
        "exam_application_use": use,
        "application_axis": axis,
        "retrieval_intent": "exam_application",
        "exam_relevance": "high",
        "is_exam_application": True,
    }


def study_priority(category_code: str | None, role: str) -> int:
    if role in {"explanation", "analysis"}:
        return 1
    if category_code in {"03", "04", "10"}:
        return 1
    if category_code in {"00", "01", "02", "05", "08"}:
        return 2
    return 3


def normalize(value: str) -> str:
    return unicodedata.normalize("NFC", value)


if __name__ == "__main__":
    raise SystemExit(main())
